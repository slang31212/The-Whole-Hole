#!/usr/bin/env python3
"""
Build a dynamic, collapsible Excel cost model for a floating offshore wind farm.

The workbook is fully formula-driven:
  * Leaf line items hold input rates (real 2024 £/MW, or £/MW/year for OPEX).
  * Every assembly / category row is a live SUM of its children, so the whole
    breakdown "rolls up" from individual line items to full wind-farm level.
  * Excel outline grouping lets each assembly collapse to its parent level.
  * A single Assumptions sheet drives capacity, life, WACC, energy yield, etc.
    Change a driver and every total, the lifetime summary, the pie chart and
    the LCOE recalculate.

Source data: BVGA-style floating offshore wind cost breakdown (real 2024 prices),
1000 MW / 15 MW turbine site. Figures are rounded, so computed roll-ups may
differ slightly from the printed category totals (shown for reference).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.worksheet.properties import Outline, PageSetupProperties

# --------------------------------------------------------------------------- #
#  Cost breakdown structure                                                     #
#  Each node: (name, stated_value, [children])                                  #
#  Leaves carry the input rate; parents are computed as SUM(children).          #
# --------------------------------------------------------------------------- #

def node(name, value, children=None):
    return {"name": name, "value": value, "children": children or []}

CAPEX = [
    node("Development and project management", 155000, [
        node("Development and consenting services", 72000, [
            node("Environmental impact assessments", 11000),
            node("Development activities and other consenting services", 62000),
        ]),
        node("Environmental surveys", 9000, [
            node("Animal surveys (benthic, fish, shellfish, mammals and birds)", 7000),
            node("Onshore environmental surveys", 1000),
            node("Human impact studies", 1000),
        ]),
        node("Resource and metocean assessment", 7000, [
            node("Structure", 4000),
            node("Sensors", 3000),
            node("Maintenance", 1000),
        ]),
        node("Geological and hydrographical surveys", 9000, [
            node("Geophysical surveys", 3000),
            node("Geotechnical surveys", 5000),
            node("Hydrographic surveys", 2000),
        ]),
        node("Engineering and consultancy", 9000),
        node("Project management", 48000),
    ]),
    node("Wind turbine", 1350000, [
        node("Nacelle", 834000),
        node("Rotor", 360000),
        node("Tower", 156000),
    ]),
    node("Balance of plant", 2418000, [
        node("Dynamic array cable", 115000),
        node("Export cable", 269000),
        node("Cable accessories", 80000, [
            node("Interface", 30000),
            node("Cable protection", 4000),
            node("Buoyancy", 2000),
            node("Connectors and joints", 43000),
        ]),
        node("Floating substructure", 1313000, [
            node("Structure", 1103000),
            node("Secondary steel", 53000),
            node("Systems", 92000),
            node("Corrosion protection", 65000),
        ]),
        node("Mooring systems", 316000, [
            node("Anchor systems", 35000),
            node("Mooring lines and chains", 174000),
            node("Jewellery", 98000),
            node("Topside connection", 6000),
            node("Installation aids", 3000),
        ]),
        node("Offshore substation", 282000, [
            node("HVAC electrical system", 80000),
            node("Auxiliary systems", 13000),
            node("Topside structure", 123000),
            node("Foundation", 65000),
        ]),
        node("Onshore substation", 44000, [
            node("Electrical system", 31000),
            node("Buildings, access and security", 13000),
        ]),
    ]),
    node("Installation and commissioning", 1376000, [
        node("Inbound transport", 154000),
        node("Mooring and anchoring pre-installation", 153000),
        node("Floating substructure - turbine assembly", 72000, [
            node("Crane and lifting equipment", 34000),
            node("Technician services", 11000),
            node("Marshalling port", 22000),
            node("Other", 5000),
        ]),
        node("Floating substructure - turbine installation", 114000),
        node("Offshore cable installation", 171000),
        node("Onshore export cable installation", 8000),
        node("Offshore substation installation", 52000),
        node("Onshore substation construction", 29000),
        node("Offshore logistics", 13000, [
            node("Sea-based support", 6000),
            node("Marine coordination", 2000),
            node("Weather forecasting and metocean data", 1000),
            node("Marine safety and rescue", 4000),
        ]),
        node("Contingency and insurance", 610000),
    ]),
]

OPEX = [
    node("Operation, maintenance and service", 98000, [
        node("Operations, maintenance and service port", 0),
        node("Operations", 34000, [
            node("Operations control centre", 1000),
            node("Training", 3000),
            node("Onshore logistics", 1000),
            node("Technical resource (onshore and off)", 7000),
            node("Admin and support staff (onshore)", 8000),
            node("Insurance", 14000),
            node("Offshore logistics", 7000),
        ]),
        node("Maintenance and service", 57000, [
            node("Turbine maintenance and service", 41000),
            node("Balance of plant maintenance and service", 15000),
            node("Statutory inspections", 1000),
        ]),
    ]),
]

DECEX = [
    node("Decommissioning", 450000, [
        node("Floating hull - turbine decommissioning", 148000),
        node("Mooring and anchoring decommissioning", 122000),
        node("Cable decommissioning", 137000),
        node("Substation decommissioning", 42000),
    ]),
]

# --------------------------------------------------------------------------- #
#  Styling helpers                                                              #
# --------------------------------------------------------------------------- #

NAVY   = "1F3B57"
OCEAN  = "2E6E8E"
STEEL  = "5B7A8C"
LIGHT  = "EAF1F5"
LIGHT2 = "F4F8FA"
ACCENT = "C8892E"
WHITE  = "FFFFFF"

thin = Side(style="thin", color="D0D8DE")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
border_bottom = Border(bottom=Side(style="thin", color=STEEL))

GBP = '£#,##0'
GBP2 = '£#,##0'
PCT = '0.0%'

LEVEL_FILL = {
    0: PatternFill("solid", fgColor=NAVY),
    1: PatternFill("solid", fgColor=LIGHT),
    2: PatternFill("solid", fgColor=LIGHT2),
    3: PatternFill("solid", fgColor=WHITE),
}
LEVEL_FONT = {
    0: Font(bold=True, color=WHITE, size=11),
    1: Font(bold=True, color=NAVY, size=11),
    2: Font(bold=False, color="1A1A1A", size=10),
    3: Font(bold=False, color="555555", size=10),
}


def build():
    wb = Workbook()

    # ------------------------------------------------------------------ #
    #  ASSUMPTIONS SHEET                                                   #
    # ------------------------------------------------------------------ #
    a = wb.active
    a.title = "Assumptions"
    a.sheet_view.showGridLines = False

    a.column_dimensions["A"].width = 3
    a.column_dimensions["B"].width = 48
    a.column_dimensions["C"].width = 16
    a.column_dimensions["D"].width = 14
    a.column_dimensions["E"].width = 60

    a["B2"] = "Floating Offshore Wind Farm — Cost Model"
    a["B2"].font = Font(bold=True, size=16, color=NAVY)
    a["B3"] = "Dynamic line-item cost breakdown, rolling up to full wind-farm level.  Real 2024 prices (£)."
    a["B3"].font = Font(italic=True, color=STEEL, size=10)

    def section(cell, text):
        a[cell] = text
        a[cell].font = Font(bold=True, color=WHITE, size=11)
        a[cell].fill = PatternFill("solid", fgColor=OCEAN)
        a[cell].alignment = Alignment(vertical="center")

    def inp(row, label, value, unit, note="", fmt=None, pct=False):
        a[f"B{row}"] = label
        a[f"B{row}"].font = Font(size=10)
        c = a[f"C{row}"]
        c.value = value
        c.font = Font(bold=True, color=ACCENT, size=11)
        c.fill = PatternFill("solid", fgColor="FBF3E6")
        c.border = border_all
        c.alignment = Alignment(horizontal="right")
        if pct:
            c.number_format = PCT
        elif fmt:
            c.number_format = fmt
        a[f"D{row}"] = unit
        a[f"D{row}"].font = Font(size=9, color=STEEL)
        if note:
            a[f"E{row}"] = note
            a[f"E{row}"].font = Font(size=9, italic=True, color=STEEL)

    section("B5", "Site definition")
    a.merge_cells("B5:E5")
    inp(6,  "Wind farm rating",                         1000, "MW",   "Total installed capacity — drives every full-farm total")
    inp(7,  "Turbine rating",                           15,   "MW",   "")
    inp(8,  "Number of turbines",                       None, "",     "= farm rating / turbine rating")
    a["C8"] = "=ROUND(C6/C7,0)"
    a["C8"].font = Font(bold=True, color=NAVY); a["C8"].number_format = "0"
    a["C8"].alignment = Alignment(horizontal="right"); a["C8"].border = border_all
    inp(9,  "Water depth at site",                      100,  "m",    "")
    inp(10, "Annual mean wind speed at 100 m",          10,   "m/s",  "")
    inp(11, "Offshore substation to shore",             75,   "km",   "")
    inp(12, "Shore to onshore substation",              10,   "km",   "")

    section("B14", "Programme")
    a.merge_cells("B14:E14")
    inp(15, "Year of FID",                              2028, "",     "Discounting base year (t = 0)")
    inp(16, "First operation date",                     2030, "",     "")
    inp(17, "Construction period",                      2,    "years","CAPEX spread across the years before first operation")
    inp(18, "Operational life",                         30,   "years","Assumption — used for OPEX lifetime and LCOE")

    section("B20", "Financial & energy")
    a.merge_cells("B20:E20")
    inp(21, "Discount rate (WACC)",                     0.08, "",     "Real, pre-tax; used for LCOE discounting", pct=True)
    inp(22, "Gross capacity factor",                    0.55, "",     "Before availability & electrical losses", pct=True)
    inp(23, "Availability",                             0.95, "",     "Technical availability of the wind farm", pct=True)
    inp(24, "Electrical / transmission losses",         0.03, "",     "Array + export + substation losses", pct=True)
    inp(25, "Net load factor",                          None, "",     "= gross CF × availability × (1 − losses)", pct=True)
    a["C25"] = "=C22*C23*(1-C24)"
    a["C25"].font = Font(bold=True, color=NAVY); a["C25"].number_format = PCT
    a["C25"].alignment = Alignment(horizontal="right"); a["C25"].border = border_all
    inp(26, "Net annual energy (AEP)",                  None, "MWh/yr","= rating × 8,760 h × net load factor")
    a["C26"] = "=C6*8760*C25"
    a["C26"].font = Font(bold=True, color=NAVY); a["C26"].number_format = "#,##0"
    a["C26"].alignment = Alignment(horizontal="right"); a["C26"].border = border_all

    a["B28"] = ("Cells shaded orange are inputs — change them and the whole model recalculates. "
                "All other figures are formulas.")
    a["B28"].font = Font(italic=True, size=9, color=STEEL)
    a.merge_cells("B28:E28")

    # Cell references used across sheets
    REF_CAP  = "Assumptions!$C$6"
    REF_LIFE = "Assumptions!$C$18"
    REF_WACC = "Assumptions!$C$21"
    REF_FID  = "Assumptions!$C$15"
    REF_FOP  = "Assumptions!$C$16"
    REF_CONS = "Assumptions!$C$17"
    REF_AEP  = "Assumptions!$C$26"

    # ------------------------------------------------------------------ #
    #  COST BREAKDOWN SHEET (the dynamic, collapsible roll-up)             #
    # ------------------------------------------------------------------ #
    cb = wb.create_sheet("Cost Breakdown")
    cb.sheet_view.showGridLines = False
    cb.sheet_properties.outlinePr = Outline(summaryBelow=False, summaryRight=False)

    widths = {"A": 3, "B": 56, "C": 12, "D": 15, "E": 18, "F": 15, "G": 13}
    for col, w in widths.items():
        cb.column_dimensions[col].width = w

    cb["B2"] = "Cost Breakdown Structure"
    cb["B2"].font = Font(bold=True, size=15, color=NAVY)
    cb["B3"] = ("Line items roll up to assemblies and full wind-farm level. "
                "Use the +/− outline buttons on the left to collapse each assembly.")
    cb["B3"].font = Font(italic=True, color=STEEL, size=10)

    header_row = 5
    headers = ["Cost item", "Unit", "Rate", "Wind farm total",
               "Stated (source)", "Variance"]
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        c = cb[f"{col}{header_row}"]
        c.value = h
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="left" if i == 0 else "right",
                                vertical="center", wrap_text=True)
        c.border = border_all
    cb.row_dimensions[header_row].height = 30

    # Sub-heading rows that separate the three cost phases
    row = header_row + 1

    def phase_header(r, text, unit_label):
        cb[f"B{r}"] = text
        cb[f"B{r}"].font = Font(bold=True, italic=True, color=OCEAN, size=10)
        cb[f"C{r}"] = unit_label
        cb[f"C{r}"].font = Font(italic=True, color=STEEL, size=9)
        cb[f"C{r}"].alignment = Alignment(horizontal="right")

    # We emit each tree, tracking rows so parents SUM their direct children.
    total_rows = {"CAPEX": [], "OPEX": [], "DECEX": []}

    def emit_tree(nodes, depth, unit, phase):
        """Recursively write rows; return list of this level's row numbers."""
        nonlocal row
        my_rows = []
        for n in nodes:
            r = row
            my_rows.append(r)
            indent = "    " * depth
            lvl = min(depth, 3)
            # Name
            bc = cb[f"B{r}"]
            bc.value = indent + n["name"]
            bc.font = LEVEL_FONT[lvl]
            bc.fill = LEVEL_FILL[lvl]
            bc.alignment = Alignment(indent=depth)
            # Unit
            uc = cb[f"C{r}"]
            uc.value = unit
            uc.font = Font(size=8, color=(WHITE if lvl == 0 else STEEL))
            uc.fill = LEVEL_FILL[lvl]
            uc.alignment = Alignment(horizontal="right")
            row += 1
            # children first (so we know their rows), then set this row's formula
            if n["children"]:
                child_rows = emit_tree(n["children"], depth + 1, unit, phase)
                # SUM of direct children only
                refs = ",".join(f"D{cr}" for cr in child_rows)
                cb[f"D{r}"] = f"=SUM({refs})"
            else:
                cb[f"D{r}"] = n["value"]
            # Wind farm total = rate * capacity
            cb[f"E{r}"] = f"=D{r}*{REF_CAP}"
            # Stated source (rounded) + variance vs computed roll-up
            cb[f"F{r}"] = n["value"]
            cb[f"G{r}"] = f"=D{r}-F{r}"

            # formats & fills for numeric cols
            for col in ("D", "E", "F", "G"):
                cell = cb[f"{col}{r}"]
                cell.fill = LEVEL_FILL[lvl]
                cell.alignment = Alignment(horizontal="right")
                if lvl == 0:
                    cell.font = Font(bold=True, color=WHITE, size=10)
                elif lvl == 1:
                    cell.font = Font(bold=True, color=NAVY, size=10)
                else:
                    cell.font = Font(color="333333", size=10)
            cb[f"D{r}"].number_format = GBP
            cb[f"E{r}"].number_format = GBP
            cb[f"F{r}"].number_format = GBP
            cb[f"G{r}"].number_format = GBP
            # % of phase total filled later (needs phase total row)
            # outline grouping
            if depth > 0:
                cb.row_dimensions[r].outline_level = min(depth, 7)

            if depth == 0:
                total_rows[phase].append(r)
        return my_rows

    # ---- CAPEX ----
    phase_header(row, "CAPEX  —  Capital expenditure (one-off)", "£/MW")
    row += 1
    emit_tree(CAPEX, 0, "£/MW", "CAPEX")
    capex_total_row = row
    cb[f"B{capex_total_row}"] = "TOTAL CAPEX (build cost)"
    refs = ",".join(f"D{r0}" for r0 in total_rows["CAPEX"])
    cb[f"D{capex_total_row}"] = f"=SUM({refs})"
    cb[f"E{capex_total_row}"] = f"=D{capex_total_row}*{REF_CAP}"
    row += 2

    # ---- OPEX ----
    phase_header(row, "OPEX  —  Operating expenditure (annual)", "£/MW/yr")
    row += 1
    emit_tree(OPEX, 0, "£/MW/yr", "OPEX")
    opex_total_row = row
    cb[f"B{opex_total_row}"] = "TOTAL OPEX (per year)"
    refs = ",".join(f"D{r0}" for r0 in total_rows["OPEX"])
    cb[f"D{opex_total_row}"] = f"=SUM({refs})"
    cb[f"E{opex_total_row}"] = f"=D{opex_total_row}*{REF_CAP}"
    row += 2

    # ---- DECEX ----
    phase_header(row, "DECEX  —  Decommissioning (end of life)", "£/MW")
    row += 1
    emit_tree(DECEX, 0, "£/MW", "DECEX")
    decex_total_row = row
    cb[f"B{decex_total_row}"] = "TOTAL DECEX (decommissioning)"
    refs = ",".join(f"D{r0}" for r0 in total_rows["DECEX"])
    cb[f"D{decex_total_row}"] = f"=SUM({refs})"
    cb[f"E{decex_total_row}"] = f"=D{decex_total_row}*{REF_CAP}"
    row += 1

    # style the three phase-total rows
    for tr in (capex_total_row, opex_total_row, decex_total_row):
        for col in ("B", "C", "D", "E", "F", "G"):
            cell = cb[f"{col}{tr}"]
            cell.fill = PatternFill("solid", fgColor=ACCENT)
            cell.font = Font(bold=True, color=WHITE, size=11)
            cell.border = Border(top=Side(style="medium", color=NAVY))
        cb[f"D{tr}"].number_format = GBP
        cb[f"E{tr}"].number_format = GBP
        cb[f"D{tr}"].alignment = Alignment(horizontal="right")
        cb[f"E{tr}"].alignment = Alignment(horizontal="right")

    # Grand lifetime totals (per MW and full farm) just below DECEX
    row += 1
    life_row = row
    cb[f"B{life_row}"] = "LIFETIME COST  (CAPEX + OPEX×life + DECEX)"
    cb[f"C{life_row}"] = "£/MW"
    cb[f"D{life_row}"] = (f"=D{capex_total_row}+D{opex_total_row}*{REF_LIFE}+D{decex_total_row}")
    cb[f"E{life_row}"] = f"=D{life_row}*{REF_CAP}"
    for col in ("B", "C", "D", "E", "F", "G"):
        cell = cb[f"{col}{life_row}"]
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(bold=True, color=WHITE, size=12)
    cb[f"D{life_row}"].number_format = GBP
    cb[f"E{life_row}"].number_format = GBP
    cb[f"C{life_row}"].alignment = Alignment(horizontal="right")
    cb[f"D{life_row}"].alignment = Alignment(horizontal="right")
    cb[f"E{life_row}"].alignment = Alignment(horizontal="right")

    # Freeze header, add autofilter over the item area
    cb.freeze_panes = "B6"

    # Note about rounding / % column
    note_row = life_row + 2
    cb[f"B{note_row}"] = ("Notes:  'Rate' leaf cells are inputs; every assembly/category is =SUM(children), "
                          "so editing a line item rolls up automatically.  'Wind farm total' = Rate × farm rating "
                          "(£/yr for OPEX rows).  'Stated (source)' is the published rounded figure; 'Variance' "
                          "is the small rounding gap between the computed roll-up and the published total.")
    cb[f"B{note_row}"].font = Font(italic=True, size=9, color=STEEL)
    cb[f"B{note_row}"].alignment = Alignment(wrap_text=True, vertical="top")
    cb.merge_cells(f"B{note_row}:G{note_row+2}")

    # ------------------------------------------------------------------ #
    #  LIFETIME SUMMARY SHEET + pie chart                                 #
    # ------------------------------------------------------------------ #
    su = wb.create_sheet("Lifetime Summary")
    su.sheet_view.showGridLines = False
    for col, w in {"A": 3, "B": 40, "C": 18, "D": 18, "E": 12}.items():
        su.column_dimensions[col].width = w

    su["B2"] = "Lifetime Cost Summary"
    su["B2"].font = Font(bold=True, size=15, color=NAVY)
    su["B3"] = "Contribution of each major cost element to whole-life cost of the wind farm."
    su["B3"].font = Font(italic=True, color=STEEL, size=10)

    hdr = 5
    for i, h in enumerate(["Cost element", "Lifetime cost (£)", "Per MW (£/MW)", "% of total"]):
        col = get_column_letter(2 + i)
        c = su[f"{col}{hdr}"]
        c.value = h
        c.font = Font(bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="left" if i == 0 else "right")
        c.border = border_all

    # rows point back to the top-level category rows on the Cost Breakdown sheet
    elements = [
        ("Development and project management", total_rows["CAPEX"][0], "CAPEX"),
        ("Wind turbine",                        total_rows["CAPEX"][1], "CAPEX"),
        ("Balance of plant",                    total_rows["CAPEX"][2], "CAPEX"),
        ("Installation and commissioning",      total_rows["CAPEX"][3], "CAPEX"),
        ("Operation, maintenance and service",  total_rows["OPEX"][0],  "OPEX"),
        ("Decommissioning",                     total_rows["DECEX"][0], "DECEX"),
    ]
    first = hdr + 1
    r = first
    for name, src_row, kind in elements:
        su[f"B{r}"] = name
        su[f"B{r}"].font = Font(size=10)
        # lifetime £ for this element
        if kind == "OPEX":
            su[f"C{r}"] = f"='Cost Breakdown'!E{src_row}*{REF_LIFE}"
        else:
            su[f"C{r}"] = f"='Cost Breakdown'!E{src_row}"
        su[f"D{r}"] = f"='Cost Breakdown'!D{src_row}" + ("" if kind != "OPEX" else f"*{REF_LIFE}")
        su[f"C{r}"].number_format = GBP
        su[f"D{r}"].number_format = GBP
        su[f"C{r}"].alignment = Alignment(horizontal="right")
        su[f"D{r}"].alignment = Alignment(horizontal="right")
        for col in ("B", "C", "D", "E"):
            su[f"{col}{r}"].border = border_all
        r += 1
    last = r - 1
    tot = r
    su[f"B{tot}"] = "TOTAL LIFETIME COST"
    su[f"C{tot}"] = f"=SUM(C{first}:C{last})"
    su[f"D{tot}"] = f"=SUM(D{first}:D{last})"
    for col, fmt in (("C", GBP), ("D", GBP)):
        su[f"{col}{tot}"].number_format = fmt
        su[f"{col}{tot}"].alignment = Alignment(horizontal="right")
    for col in ("B", "C", "D", "E"):
        su[f"{col}{tot}"].fill = PatternFill("solid", fgColor=ACCENT)
        su[f"{col}{tot}"].font = Font(bold=True, color=WHITE)
    # percentages
    for rr in range(first, last + 1):
        su[f"E{rr}"] = f"=C{rr}/$C${tot}"
        su[f"E{rr}"].number_format = PCT
        su[f"E{rr}"].alignment = Alignment(horizontal="right")
    su[f"E{tot}"] = f"=SUM(E{first}:E{last})"
    su[f"E{tot}"].number_format = PCT
    su[f"E{tot}"].alignment = Alignment(horizontal="right")

    # Pie chart of lifetime contribution
    pie = PieChart()
    pie.title = "Lifetime cost contribution by element"
    labels = Reference(su, min_col=2, min_row=first, max_row=last)
    data = Reference(su, min_col=3, min_row=hdr, max_row=last)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.height = 9
    pie.width = 16
    su.add_chart(pie, f"B{tot + 3}")

    # ------------------------------------------------------------------ #
    #  LCOE SHEET — discounted cash-flow model                            #
    # ------------------------------------------------------------------ #
    lc = wb.create_sheet("LCOE")
    lc.sheet_view.showGridLines = False
    for col, w in {"A": 3, "B": 10, "C": 8, "D": 18, "E": 18, "F": 18,
                   "G": 16, "H": 16, "I": 18, "J": 18}.items():
        lc.column_dimensions[col].width = w

    lc["B2"] = "Levelised Cost of Energy (LCOE)"
    lc["B2"].font = Font(bold=True, size=15, color=NAVY)
    lc["B3"] = ("Discounted cash-flow model.  LCOE = Σ(discounted cost) / Σ(discounted energy), "
                "discounted to the year of FID.")
    lc["B3"].font = Font(italic=True, color=STEEL, size=10)

    # headline result box
    lc["B5"] = "LCOE"
    lc["B5"].font = Font(bold=True, color=NAVY, size=12)
    # (formula added after table)

    thr = 8
    lheaders = ["Year", "t", "CAPEX (£)", "OPEX (£)", "DECEX (£)",
                "Total cost (£)", "Energy (MWh)", "Discount factor",
                "Disc. cost (£)", "Disc. energy (MWh)"]
    for i, h in enumerate(lheaders):
        col = get_column_letter(2 + i)
        c = lc[f"{col}{thr}"]
        c.value = h
        c.font = Font(bold=True, color=WHITE, size=9)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border_all
    lc.row_dimensions[thr].height = 30

    # Build a year-by-year timeline: FID .. first_operation-1 (construction),
    # then Life operational years, then 1 decommissioning year.
    # We hard-size to a generous 40 rows and drive contents by formulas so the
    # user can change Life / dates and it stays consistent.
    n_rows = 45
    start = thr + 1
    capex_full = f"'Cost Breakdown'!$E${capex_total_row}"
    opex_full  = f"'Cost Breakdown'!$E${opex_total_row}"
    decex_full = f"'Cost Breakdown'!$E${decex_total_row}"

    for k in range(n_rows):
        r = start + k
        # Year
        lc[f"B{r}"] = f"={REF_FID}+{k}"
        lc[f"B{r}"].number_format = "0"
        # t (years from FID)
        lc[f"C{r}"] = k
        # CAPEX: spread equally across construction years (FID .. first op - 1)
        lc[f"D{r}"] = (f"=IF(B{r}<{REF_FOP},{capex_full}/{REF_CONS},0)")
        # OPEX: during operational years (first op .. first op + life - 1)
        lc[f"E{r}"] = (f"=IF(AND(B{r}>={REF_FOP},B{r}<{REF_FOP}+{REF_LIFE}),{opex_full},0)")
        # DECEX: the year after the last operational year
        lc[f"F{r}"] = (f"=IF(B{r}={REF_FOP}+{REF_LIFE},{decex_full},0)")
        # Total cost
        lc[f"G{r}"] = f"=D{r}+E{r}+F{r}"
        # Energy: during operational years
        lc[f"H{r}"] = (f"=IF(AND(B{r}>={REF_FOP},B{r}<{REF_FOP}+{REF_LIFE}),{REF_AEP},0)")
        # Discount factor to FID
        lc[f"I{r}"] = f"=1/(1+{REF_WACC})^C{r}"
        lc[f"I{r}"].number_format = "0.0000"
        # Discounted cost / energy
        lc[f"J{r}"] = f"=G{r}*I{r}"
        lc[f"K{r}"] = f"=H{r}*I{r}"
        for col in ("D", "E", "F", "G", "J"):
            lc[f"{col}{r}"].number_format = GBP
        lc[f"H{r}"].number_format = "#,##0"
        lc[f"K{r}"].number_format = "#,##0"
        for col in "BCDEFGHIJK":
            lc[f"{col}{r}"].border = border_all
            if k % 2:
                lc[f"{col}{r}"].fill = PatternFill("solid", fgColor=LIGHT2)

    end = start + n_rows - 1
    tr = end + 1
    lc[f"B{tr}"] = "Totals"
    lc[f"B{tr}"].font = Font(bold=True, color=WHITE)
    for col in ("D", "E", "F", "G", "H", "J", "K"):
        lc[f"{col}{tr}"] = f"=SUM({col}{start}:{col}{end})"
        lc[f"{col}{tr}"].number_format = GBP if col in ("D","E","F","G","J") else "#,##0"
    for col in "BCDEFGHIJK":
        lc[f"{col}{tr}"].fill = PatternFill("solid", fgColor=ACCENT)
        lc[f"{col}{tr}"].font = Font(bold=True, color=WHITE)

    # LCOE headline = total discounted cost / total discounted energy
    lc["C5"] = f"=J{tr}/K{tr}"
    lc["C5"].number_format = '£#,##0.00" /MWh"'
    lc["C5"].font = Font(bold=True, color=WHITE, size=14)
    lc["C5"].fill = PatternFill("solid", fgColor=OCEAN)
    lc["C5"].alignment = Alignment(horizontal="center")
    lc.merge_cells("C5:E5")
    lc["B6"] = "£ per MWh, real 2024, discounted to FID"
    lc["B6"].font = Font(italic=True, size=9, color=STEEL)
    lc.merge_cells("B6:E6")

    lc.freeze_panes = f"B{start}"

    # ------------------------------------------------------------------ #
    wb.save("Floating_Offshore_Wind_Farm_Cost_Model.xlsx")
    print("Saved Floating_Offshore_Wind_Farm_Cost_Model.xlsx")
    print(f"  CAPEX total row: {capex_total_row}, OPEX: {opex_total_row}, DECEX: {decex_total_row}")


if __name__ == "__main__":
    build()
